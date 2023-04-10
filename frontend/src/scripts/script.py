
# -*- coding: utf-8 -*-
"""
Created on Fri Mar 31 17:40:27 2023

@author: const
"""

import openpyxl
import pandas as pd


def genereFeuille(filter_type, filter_value=None):
    # On charge les données de "BDD.xlsx"
    data = pd.read_excel("BDD.xlsx")

    # On filtre les données en fonction du filtre choisi
    if filter_type == "MAISON":
        # enlève les doublons et transforme en liste
        houses = data['maison'].unique().tolist()
        filtered_data = pd.DataFrame()
        for house in houses:
            players = data[data['maison'] == house]
            filtered_data = pd.concat([filtered_data, players])
    elif filter_type == "ELO":
        filtered_data = data[(data['ELO'] >= filter_value[0]) & (data['ELO'] <= filter_value[1])]
    elif filter_type == "CLASSE":
        filtered_data = data[data['classe'] == filter_value]
    else:
        filtered_data = data

     # Tri des joueurs par ordre alphabétique
    filtered_data = filtered_data.sort_values(by=['nom'])

    # Création des paires de joueurs avec l'algorithme du Round Robin
    num_players = len(filtered_data)
    if num_players % 2 == 1:
        # Ajout d'un joueur fictif pour que le nombre de joueurs soit pair
        filtered_data = pd.concat([filtered_data, pd.DataFrame({'nom': ['Joueur Fantôme'], 'maison': [''], 'ELO': [-1], 'classe': ['']})], ignore_index=True)
        num_players += 1
    half_num_players = num_players // 2
    player_list = list(filtered_data.index)
    rounds = []
    for i in range(num_players - 1):
        round = []
        for j in range(half_num_players):
            if player_list[j] != None and player_list[num_players - 1 - j] != None:
                # On ne prend pas en compte le joueur fictif "Joueur Fantôme" lors de la création des paires
                if "Joueur Fantôme" not in [filtered_data.loc[player_list[j]]['nom'], filtered_data.loc[player_list[num_players - 1 - j]]['nom']]:
                    # Deux joueurs de la même maison ne peuvent pas s'affronter
                    if filtered_data.loc[player_list[j]]['maison'] != filtered_data.loc[player_list[num_players - 1 - j]]['maison']:
                        round.append((player_list[j], player_list[num_players - 1 - j]))
        rounds.append(round)
        player_list.insert(1, player_list.pop())

    # Écriture des paires de joueurs dans la feuille de tournoi
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.append(["N° ronde", "Joueur 1", "Joueur 2", "Gagnant du match"])

    # Redimension des colonnes Excel
    worksheet.column_dimensions['A'].width = 10
    worksheet.column_dimensions['B'].width = 13
    worksheet.column_dimensions['C'].width = 13
    worksheet.column_dimensions['D'].width = 17

    for round_num, round in enumerate(rounds):
        for pair in round:
            player1 = filtered_data.loc[pair[0]]
            player2 = filtered_data.loc[pair[1]]
            worksheet.append([f"Ronde {round_num + 1}", f"{player1['nom']}", f"{player2['nom']}", None])

            #players_str = f"{player1['nom']} {player1['prénom']} vs {player2['nom']} {player2['prénom
            #worksheet.append([f"Ronde {round_num + 1}", player1['nom'], player1['prénom'], player2['nom'], player2['prénom'], None])


    # Définition du nom du fichier Excel en fonction du filtre choisi et de la date du jour
    date_format = '%d-%m-%Y'
    date_str = pd.Timestamp.now().strftime(date_format)
    if filter_type == "MAISON":
        filename = f"Tournoi par Maison ({date_str}).xlsx"
    elif filter_type == "ELO":
        filename = f"Tournoi par ELO ({filter_value[0]}-{filter_value[1]}) ({date_str}).xlsx"
    elif filter_type == "CLASSE":
        filename = f"Tournoi par Classe ({filter_value}) ({date_str}).xlsx"
    else:
        filename = f"Tournoi général ({date_str}).xlsx"

    # Enregistrement du fichier Excel
    print("feuille de tournoi générée")
    workbook.save(filename)


def expectedScore(R):
    x = 1/(1+pow(10,(R)/400)) ## ce nombre est à quelle différence d'elo on a environ 100% de victoire
    return x

#entrer 1 pour victoire J1, 0 pour défaite J1, et 2 pour match nul
def calculElo(elo1, elo2, resultatMatch):

    if(resultatMatch == 2): resultatMatch=1/2

    expected1 = expectedScore(elo2 - elo1)
    K = 40
    elo1 = elo1 + round(K*(resultatMatch - expected1),1)
    elo2 = elo2 + round(K*(expected1 - resultatMatch),1)

    return int(elo1), int(elo2)

#calculElo(544, 554, 0)



#fonction qui recupère l'elo dans BDD
def getElo(bdd_file):
    # Charger le fichier Excel en tant que DataFrame
    df = pd.read_excel(bdd_file)
    # Créer un dictionnaire pour stocker l'Elo de chaque joueur
    elos = {}
    # Parcourir chaque ligne du DataFrame
    for index, row in df.iterrows():
        # Lire le nom du joueur et son Elo
        nom, elo = row['nom'], row['ELO']
        # Stocker l'Elo dans le dictionnaire
        elos[nom] = elo
    # Afficher l'Elo de chaque joueur
    for nom, elo in elos.items():
        print(f"L'Elo actuel de {nom} est {elo}.")
    return elos


#fonction qui compte le nombre de victoires d'un joueur / nb de parties effectuées 
#pendant un tournoi et affiche un pourcentage de victoire

def compteVictoire(tournoi_file):    
    # Charger le fichier Excel en tant que DataFrame
    df = pd.read_excel(tournoi_file)
    # Créer des dictionnaires pour stocker le nombre de victoires et de matchs joués de chaque joueur
    victoires = {}
    parties = {} 
    matchs_nuls = {}
    # Parcourir chaque ligne du DataFrame
    for index, row in df.iterrows():
        # Lire les noms des joueurs et le nom du gagnant
        joueur1, joueur2, gagnant = row['Joueur 1'], row['Joueur 2'], row['Gagnant du match']
        # Mettre à jour le nombre de matchs joués de chaque joueur
        parties[joueur1] = parties.get(joueur1, 0) + 1
        parties[joueur2] = parties.get(joueur2, 0) + 1
        # Mettre à jour le nombre de victoires de chaque joueur
        if gagnant == joueur1:
            victoires[joueur1] = victoires.get(joueur1, 0) + 1
        elif gagnant == joueur2:
            victoires[joueur2] = victoires.get(joueur2, 0) + 1
        elif pd.isna(gagnant):  # Condition pour les matchs nuls
            matchs_nuls[joueur1] = matchs_nuls.get(joueur1, 0) + 1
            matchs_nuls[joueur2] = matchs_nuls.get(joueur2, 0) + 1
    # Afficher le nombre de victoires et de matchs joués de chaque joueur
    #for joueur, nb_victoires in victoires.items():
    joueurs = set(victoires.keys()) | set(parties.keys())
    for joueur in joueurs:
       nb_victoires = victoires.get(joueur, 0)
       nb_parties = parties.get(joueur, 0)
       nb_matchs_nuls = matchs_nuls.get(joueur, 0)
       pourcentage_victoires = nb_victoires / nb_parties * 100
       print(f"{joueur} a gagné {nb_victoires} matchs dont {nb_matchs_nuls} nuls, sur un total de {nb_parties} parties ({pourcentage_victoires:.2f}%).")
    return victoires, parties



def updateElo(bdd_file, tournoi_file): 
    # On charge les données du "BDD.xlsx" et de "Tournoi.xlsx"
    bdd_df = pd.read_excel(bdd_file)
    tournoi_df = pd.read_excel(tournoi_file)

    # Ajout de la colonne 'ancien elo' dans la BDD pour stocker les anciennes valeurs
    bdd_df = bdd_df.assign(ancien_elo=bdd_df['ELO'])

    # Mise à jour des Elo des joueurs en parcourant les rounds du tournoi
    for i in range(len(tournoi_df)):
        player1 = tournoi_df.loc[i, 'Joueur 1']
        player2 = tournoi_df.loc[i, 'Joueur 2']
        winner = tournoi_df.loc[i, 'Gagnant du match']
           
        # Récupération des valeurs d'Elo dans bdd_df
        player1_elo = bdd_df.loc[(bdd_df['nom'] == player1), 'ELO'].item()
        player2_elo = bdd_df.loc[(bdd_df['nom'] == player2), 'ELO'].item()
                
        # Calcul du nouvel Elo pour les joueurs et mise à jour des valeurs d'Elo dans bdd_df
        if winner == player1:
            winner_elo, loser_elo = calculElo(player1_elo, player2_elo, 1)            
            bdd_df.loc[(bdd_df['nom'] == player1), 'ELO'] = winner_elo
            bdd_df.loc[(bdd_df['nom'] == player2), 'ELO'] = loser_elo
            
        elif winner == player2:
            winner_elo, loser_elo = calculElo(player2_elo, player1_elo, 1)
            bdd_df.loc[(bdd_df['nom'] == player2), 'ELO'] = winner_elo
            bdd_df.loc[(bdd_df['nom'] == player1), 'ELO'] = loser_elo
            
        elif pd.isna(winner):  # Condition pour les matchs nuls    
            winner_elo, loser_elo = calculElo(player1_elo, player2_elo, 2)            
            bdd_df.loc[(bdd_df['nom'] == player1), 'ELO'] = winner_elo
            bdd_df.loc[(bdd_df['nom'] == player2), 'ELO'] = loser_elo
            
    # Écriture dans le fichier "BDD.xlsx" avec la colonne 'ELO' mise à jour
    bdd_df.to_excel(bdd_file, index=False)   
    redimension("BDD.xlsx")
    

#fonction qui ajoute le nombre de victoires et de parties jouées au Excel
def updateNb(bdd_file, tournoi_file):
    # Calculer le nombre de victoires et de matchs joués de chaque joueur dans le tournoi
    victoires, parties = compteVictoire(tournoi_file)
    # Charger le fichier Excel en tant que DataFrame
    df = pd.read_excel(bdd_file)
    # Ajouter une colonne "nb victoires" au DataFrame
    df['nb victoires'] = df['nom'].map(victoires).fillna(0)
    # Ajouter une colonne "nb parties jouées" au DataFrame
    df['nb parties jouées'] = df['nom'].map(parties).fillna(0)
    # Sauvegarder le DataFrame dans le fichier Excel
    df.to_excel(bdd_file, index=False)



def redimension(bdd_file):
    # Ouvrir le fichier Excel avec openpyxl
    wb = openpyxl.load_workbook(bdd_file)
    # Pour chaque feuille dans le fichier Excel
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        # Pour chaque colonne dans la feuille
        for col in sheet.columns:
            # Déterminer la longueur maximale de chaque cellule dans la colonne
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Redimensionner la colonne avec la longueur maximale
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width
    # Enregistrer les modifications dans le fichier Excel
    wb.save(bdd_file)


def completeBDD(bdd, tournoi):
    updateElo(bdd, tournoi)
    print("\n")
    getElo(bdd)
    print("\n")
    updateNb(bdd, tournoi)
    redimension(bdd)
    redimension(tournoi)
    
#on génère la feuille de tournoi en fonction de nos filtres 
#genereFeuille("ELO", (400, 800)) 
 
#puis après l'avoir remplie avec le nom des gagnants, on complète la BDD à l'issue du tournoi
bdd = "BDD.xlsx"
tournoi = "tournoi.xlsx"
completeBDD(bdd,tournoi)

